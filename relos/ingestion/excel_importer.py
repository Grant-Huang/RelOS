"""
relos/ingestion/excel_importer.py
----------------------------------
Excel 批量导入 Pipeline。

将历史告警/关系数据从 Excel 文件批量转换为 RelationObject，
并写入 Neo4j 图数据库。

列名映射配置（docs/data-model.md §Excel 导入规范）：
    source_node_id    → 起始节点 ID（必填）
    source_node_type  → 起始节点类型（必填）
    target_node_id    → 目标节点 ID（必填）
    target_node_type  → 目标节点类型（必填）
    relation_type     → 关系类型（必填）
    confidence        → 初始置信度，缺省 0.75
    provenance        → 来源类型，缺省 mes_structured
    provenance_detail → 来源详情（可选）
    extracted_by      → 录入人 ID（可选）
    half_life_days    → 半衰期（可选）

设计：
- 行级校验：每行独立处理，单行失败不影响其他行
- 导入结果汇总：返回成功/失败计数及详细错误
- 来源默认为 mes_structured（Excel 属于结构化导入）
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any

import structlog
from pydantic import ValidationError

from relos.core.models import RelationObject, SourceType

logger = structlog.get_logger(__name__)

# ─── 列名映射（Excel 列名 → RelationObject 字段）────────────────────

# 标准列名（用户 Excel 中期望的中文/英文列名 → 内部字段名）
COLUMN_MAPPING: dict[str, str] = {
    # 英文列名
    "source_node_id":    "source_node_id",
    "source_node_type":  "source_node_type",
    "target_node_id":    "target_node_id",
    "target_node_type":  "target_node_type",
    "relation_type":     "relation_type",
    "confidence":        "confidence",
    "provenance":        "provenance",
    "provenance_detail": "provenance_detail",
    "extracted_by":      "extracted_by",
    "half_life_days":    "half_life_days",
    # 中文列名映射
    "起始节点ID":         "source_node_id",
    "起始节点类型":        "source_node_type",
    "目标节点ID":         "target_node_id",
    "目标节点类型":        "target_node_type",
    "关系类型":           "relation_type",
    "置信度":            "confidence",
    "来源类型":           "provenance",
    "来源详情":           "provenance_detail",
    "录入人":            "extracted_by",
    "半衰期(天)":        "half_life_days",
}

REQUIRED_FIELDS = {
    "source_node_id",
    "source_node_type",
    "target_node_id",
    "target_node_type",
    "relation_type",
}


# ─── 结果数据类 ───────────────────────────────────────────────────────

@dataclass
class RowError:
    row_number: int
    raw_data: dict[str, Any]
    error: str


@dataclass
class ImportResult:
    """Excel 导入的汇总结果。"""
    total_rows: int = 0
    success_count: int = 0
    failed_count: int = 0
    relations: list[RelationObject] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    @property
    def accuracy(self) -> float:
        """导入准确率（成功行数 / 总行数）。"""
        if self.total_rows == 0:
            return 0.0
        return round(self.success_count / self.total_rows, 4)

    def summary(self) -> dict[str, Any]:
        return {
            "total_rows": self.total_rows,
            "success_count": self.success_count,
            "failed_count": self.failed_count,
            "accuracy": self.accuracy,
            "errors": [
                {"row": e.row_number, "error": e.error}
                for e in self.errors
            ],
        }


# ─── 核心解析逻辑 ─────────────────────────────────────────────────────

class ExcelImporter:
    """
    Excel 批量导入器。

    Usage:
        importer = ExcelImporter()
        result = importer.parse_file("/path/to/relations.xlsx")
        # result.relations 是 list[RelationObject]
    """

    def __init__(
        self,
        default_provenance: SourceType = SourceType.MES_STRUCTURED,
        default_confidence: float = 0.75,
        default_half_life: int = 90,
    ) -> None:
        self.default_provenance = default_provenance
        self.default_confidence = default_confidence
        self.default_half_life = default_half_life

    def parse_file(self, file_path: str, sheet_name: str = 0) -> ImportResult:  # type: ignore[assignment]
        """
        从磁盘文件解析 Excel。

        Args:
            file_path: Excel 文件路径（.xlsx 或 .xls）
            sheet_name: 工作表名称或索引，默认第一张

        Returns:
            ImportResult 汇总
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Excel 导入需要 openpyxl 库，请运行: pip install openpyxl>=3.1.0"
            )

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        return self._parse_rows(rows)

    def parse_bytes(self, content: bytes, sheet_name: str = 0) -> ImportResult:  # type: ignore[assignment]
        """
        从字节流（例如 HTTP 上传）解析 Excel。
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "Excel 导入需要 openpyxl 库，请运行: pip install openpyxl>=3.1.0"
            )

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        return self._parse_rows(rows)

    def _parse_rows(self, rows: list[tuple[Any, ...]]) -> ImportResult:
        """将原始行数据解析为 RelationObject 列表。"""
        result = ImportResult()

        if not rows:
            logger.warning("excel_import_empty_file")
            return result

        # 第一行为表头
        header_row = rows[0]
        col_map = self._build_col_map(header_row)

        if not col_map:
            raise ValueError(
                f"无法识别 Excel 列名。期望列名包含：{', '.join(REQUIRED_FIELDS)}"
            )

        # 检查必填列
        missing_required = REQUIRED_FIELDS - set(col_map.values())
        if missing_required:
            raise ValueError(f"Excel 缺少必填列：{', '.join(missing_required)}")

        data_rows = rows[1:]
        result.total_rows = len(data_rows)

        for row_idx, row in enumerate(data_rows, start=2):  # 行号从 2 开始（1 是表头）
            # 跳过完全空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                result.total_rows -= 1
                continue

            try:
                relation = self._parse_row(row, col_map, row_idx)
                result.relations.append(relation)
                result.success_count += 1
            except (ValueError, ValidationError) as exc:
                result.failed_count += 1
                raw: dict[str, Any] = {
                    str(header_row[i]): row[i]
                    for i in range(min(len(header_row), len(row)))
                    if header_row[i] is not None
                }
                result.errors.append(RowError(
                    row_number=row_idx,
                    raw_data=raw,
                    error=str(exc),
                ))
                logger.warning(
                    "excel_row_parse_failed",
                    row=row_idx,
                    error=str(exc)[:200],
                )

        logger.info(
            "excel_import_complete",
            total=result.total_rows,
            success=result.success_count,
            failed=result.failed_count,
            accuracy=result.accuracy,
        )
        return result

    def _build_col_map(self, header_row: tuple[Any, ...]) -> dict[int, str]:
        """
        将 Excel 表头行映射到内部字段名。
        返回 {列索引: 内部字段名}。
        """
        col_map: dict[int, str] = {}
        for col_idx, cell_value in enumerate(header_row):
            if cell_value is None:
                continue
            col_name = str(cell_value).strip()
            internal_name = COLUMN_MAPPING.get(col_name)
            if internal_name:
                col_map[col_idx] = internal_name
        return col_map

    def _parse_row(
        self,
        row: tuple[Any, ...],
        col_map: dict[int, str],
        row_number: int,
    ) -> RelationObject:
        """将单行数据转换为 RelationObject。"""
        data: dict[str, Any] = {}
        for col_idx, field_name in col_map.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None and str(value).strip() != "":
                    data[field_name] = value

        # 设置默认值
        data.setdefault("provenance", self.default_provenance.value)
        data.setdefault("confidence", self.default_confidence)
        data.setdefault("half_life_days", self.default_half_life)

        # 数值类型转换
        if "confidence" in data:
            data["confidence"] = float(data["confidence"])
        if "half_life_days" in data:
            data["half_life_days"] = int(data["half_life_days"])

        # relation_type 大写规范化
        if "relation_type" in data:
            data["relation_type"] = str(data["relation_type"]).strip().upper()

        # 节点 ID 字符串化
        for id_field in ("source_node_id", "target_node_id"):
            if id_field in data:
                data[id_field] = str(data[id_field]).strip()

        # 构建 RelationObject（会触发 Pydantic 校验）
        return RelationObject(**data)
